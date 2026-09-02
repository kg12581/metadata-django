import json
import subprocess
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

from django.http import HttpResponse, JsonResponse
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from .config import get_datax_config, get_database_config, get_doris_config
from .models import (
    AnalyticsEvent,
    LineageEdge,
    MetadataDatabase,
    MetadataSourceConfig,
    MetadataTable,
    ReconcileRun,
    ReconcileTask,
    SourceType,
)
from .serializers import database_to_dict, source_config_to_dict, table_to_dict
from .services.datax_sync import sync_table_data
from .services.flink_sync import apply_job, check_all_jobs, generate_job
from .services import docs as docs_service
from .services import lineage as lineage_service
from .services import llm as llm_service
from .services import sql_files as sql_files_service
from .services import scripts as scripts_service
from .services.reconcile_engine import run_task as run_reconcile_task
from .services.schema_check import check_tables
from .services.schema_sync import sync_table_schema
from .services.sql_helper import build_snippets
from .services.sync import sync_metadata
from .services.task_config import (
    cron_state,
    install_cron,
    load_task,
    remove_cron,
    save_task,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
FLINK_SQL_DIR = PROJECT_ROOT / "flink_sql"


def _ok(data=None, message="ok", code=0, status=200):
    return JsonResponse({"code": code, "message": message, "data": data}, status=status)


def _fail(message, code=500, status=500):
    return JsonResponse({"code": code, "message": message, "data": None}, status=status)


def _parse_json_body(request) -> tuple[dict | None, str | None]:
    if not request.body:
        return {}, None
    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return None, "请求体不是合法的 JSON"
    if not isinstance(data, dict):
        return None, "请求体必须是 JSON 对象"
    return data, None


def _tables_from_payload(payload: dict) -> tuple[list[str] | None, str | None]:
    if isinstance(payload.get("table"), str):
        return [payload["table"]], None
    tables = payload.get("tables")
    if isinstance(tables, list) and tables and all(isinstance(t, str) and t for t in tables):
        return tables, None
    return None, "缺少 table 或 tables 参数"


@require_GET
def database_list(request):
    databases = MetadataDatabase.objects.prefetch_related("tables").all()
    return _ok([database_to_dict(db) for db in databases])


@require_GET
def database_detail(request, pk):
    database = get_object_or_404(
        MetadataDatabase.objects.prefetch_related("tables__columns"), pk=pk
    )
    return _ok(database_to_dict(database, include_tables=True))


@require_GET
def table_detail(request, pk):
    table = get_object_or_404(
        MetadataTable.objects.select_related("database").prefetch_related(
            "columns", "indexes", "constraints"
        ),
        pk=pk,
    )
    return _ok(table_to_dict(table, include_children=True))


@csrf_exempt
@require_POST
def sync_database(request):
    """POST /api/metadata/sync/ 触发同步, 请求体可覆盖连接配置。"""
    overrides = {}
    if request.body:
        try:
            overrides = json.loads(request.body.decode("utf-8") or "{}")
        except json.JSONDecodeError:
            return _fail("请求体不是合法的 JSON", code=400, status=400)
        if not isinstance(overrides, dict):
            return _fail("请求体必须是 JSON 对象", code=400, status=400)

    try:
        config = get_database_config(overrides)
    except (ValueError, TypeError) as exc:
        return _fail(str(exc), code=400, status=400)

    try:
        database, stats = sync_metadata(config)
    except Exception as exc:  # 连接/查询失败
        return _fail(f"同步失败: {exc}", code=500, status=500)

    return _ok(
        {
            "database": database_to_dict(database),
            "stats": stats,
        },
        message=f"同步完成, 共 {stats['tables']} 张表",
    )


@require_GET
def dashboard(request):
    """前端页面: 数据源总览。"""
    databases = (
        MetadataDatabase.objects.annotate(table_count=Count("tables"))
        .order_by("-updated_at")
    )
    return render(request, "common/dashboard.html", {"databases": databases})


@require_GET
def database_detail_ui(request, pk):
    """前端页面: 某个数据源下的表列表。"""
    database = get_object_or_404(
        MetadataDatabase.objects.annotate(table_count=Count("tables")),
        pk=pk,
    )
    tables = (
        database.tables.annotate(column_count=Count("columns"))
        .order_by("schema_name", "name")
    )
    return render(
        request,
        "common/database_detail.html",
        {"database": database, "tables": tables},
    )


@require_GET
def table_detail_ui(request, pk):
    """前端页面: 表详情(字段/索引/约束)。"""
    table = get_object_or_404(
        MetadataTable.objects.select_related("database").prefetch_related(
            "columns", "indexes", "constraints"
        ),
        pk=pk,
    )
    return render(request, "common/table_detail.html", {"table": table})


@require_GET
def export_database_excel(request, pk):
    """把某个数据源的全部元数据导出为 Excel (.xlsx)。"""
    database = get_object_or_404(
        MetadataDatabase.objects.prefetch_related(
            "tables__columns", "tables__indexes", "tables__constraints"
        ),
        pk=pk,
    )
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "缺少 openpyxl, 请先执行: pip install -r requirements.txt"
        ) from exc

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center")

    def write_sheet(title: str, headers: list[str], rows: list[list], widths: list[int]) -> None:
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        for row in rows:
            ws.append(row)
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    tables = list(database.tables.order_by("schema_name", "name"))

    table_rows = [
        [t.schema_name, t.name, t.table_type, t.comment, t.columns.count()]
        for t in tables
    ]
    write_sheet(
        "表",
        ["Schema", "表名", "类型", "注释", "字段数"],
        table_rows,
        [18, 30, 14, 40, 10],
    )

    column_rows = [
        [
            c.table.schema_name,
            c.table.name,
            c.ordinal_position,
            c.name,
            c.data_type,
            c.column_type,
            "是" if c.is_nullable else "否",
            c.column_default or "",
            c.comment,
        ]
        for t in tables
        for c in t.columns.all()
    ]
    write_sheet(
        "字段",
        ["Schema", "表名", "序号", "字段名", "数据类型", "完整类型", "可空", "默认值", "注释"],
        column_rows,
        [18, 30, 8, 25, 18, 18, 8, 30, 40],
    )

    index_rows = [
        [
            i.table.schema_name,
            i.table.name,
            i.name,
            "主键" if i.is_primary else ("唯一" if i.is_unique else "普通"),
            ", ".join(i.column_names),
            i.definition,
        ]
        for t in tables
        for i in t.indexes.all()
    ]
    write_sheet(
        "索引",
        ["Schema", "表名", "索引名", "类型", "字段", "定义"],
        index_rows,
        [18, 30, 30, 10, 40, 60],
    )

    constraint_rows = [
        [
            c.table.schema_name,
            c.table.name,
            c.name,
            c.constraint_type,
            ", ".join(c.column_names),
            c.referenced_table,
            c.referenced_column,
        ]
        for t in tables
        for c in t.constraints.all()
    ]
    write_sheet(
        "约束",
        ["Schema", "表名", "约束名", "类型", "字段", "引用表", "引用字段"],
        constraint_rows,
        [18, 30, 30, 16, 40, 30, 30],
    )

    del wb["Sheet"]  # 删除默认空 sheet

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"metadata_{database.database_name}_{timezone.now():%Y%m%d_%H%M%S}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@csrf_exempt
@require_POST
def datax_check(request):
    """校验 MySQL 与 Doris 表结构是否一致, 返回 data.consistent = true/false。"""
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)

    mysql_config = get_database_config(payload)
    if mysql_config["db_type"] != "mysql":
        return _fail("结构校验仅支持 MySQL 作为源库", code=400, status=400)
    tables, err = _tables_from_payload(payload)
    if err:
        return _fail(err, code=400, status=400)

    doris_config = get_doris_config(payload)
    result = check_tables(
        mysql_config,
        doris_config,
        mysql_config["database"],
        tables,
        doris_database=payload.get("doris_database"),
    )
    message = "表结构一致" if result["consistent"] else "存在表结构差异"
    return _ok(result, message=message)


@csrf_exempt
@require_POST
def datax_sync(request):
    """先校验 MySQL 与 Doris 表结构一致, 再执行 DataX 同步表数据。"""
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)

    mysql_config = get_database_config(payload)
    if mysql_config["db_type"] != "mysql":
        return _fail("DataX 同步仅支持 MySQL 作为源库", code=400, status=400)
    tables, err = _tables_from_payload(payload)
    if err:
        return _fail(err, code=400, status=400)

    doris_config = get_doris_config(payload)
    datax_config = get_datax_config(payload)
    doris_database = payload.get("doris_database")
    truncate = bool(payload.get("truncate", True))
    channel = int(payload.get("channel", 3))
    force = bool(payload.get("force", False))
    preview = bool(payload.get("preview", False))
    split_pk = payload.get("split_pk") or None

    check = check_tables(
        mysql_config,
        doris_config,
        mysql_config["database"],
        tables,
        doris_database=doris_database,
    )
    if not check["consistent"] and not force and not preview:
        return JsonResponse(
            {
                "code": 409,
                "message": "表结构不一致, 未执行 DataX 同步(如需强制执行请传 force=true)",
                "data": check,
            },
            status=409,
        )

    results = []
    for table in tables:
        try:
            result = sync_table_data(
                mysql_config,
                doris_config,
                datax_config,
                mysql_config["database"],
                table,
                doris_database=doris_database,
                truncate=truncate,
                channel=channel,
                force=force,
                split_pk=split_pk,
                preview=preview,
            )
        except Exception as exc:
            result = {
                "table": table,
                "checked": False,
                "consistent": False,
                "executed": False,
                "success": False,
                "reason": str(exc),
            }
        results.append(result)

    executed = [r for r in results if r.get("executed")]
    ok = all(r.get("success") for r in executed)
    if preview:
        message = f"preview 模式, 已生成 {len(results)} 个 DataX job"
    elif ok:
        message = f"DataX 同步完成, 共 {len(executed)} 张表"
    else:
        message = "DataX 同步存在失败, 详见 results"
    return _ok({"check": check, "results": results}, message=message)


@csrf_exempt
@require_POST
def schema_sync(request):
    """根据 MySQL 元数据自动对齐 Doris 表结构(新增/删除/修改字段, 不存在自动建表)。

    请求体: {database, table|tables, doris_database, preview=true(默认), drop_columns=true, auto_create=true}
    """
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)

    mysql_config = get_database_config(payload)
    if mysql_config["db_type"] != "mysql":
        return _fail("结构同步仅支持 MySQL 作为源库", code=400, status=400)
    tables, err = _tables_from_payload(payload)
    if err:
        return _fail(err, code=400, status=400)

    doris_config = get_doris_config(payload)
    preview = bool(payload.get("preview", True))
    drop_columns = bool(payload.get("drop_columns", True))
    auto_create = bool(payload.get("auto_create", True))

    results = []
    for table in tables:
        try:
            result = sync_table_schema(
                mysql_config,
                doris_config,
                mysql_config["database"],
                table,
                doris_database=payload.get("doris_database"),
                preview=preview,
                drop_columns=drop_columns,
                auto_create=auto_create,
            )
        except Exception as exc:
            result = {"table": table, "error": str(exc), "statements": [], "warnings": []}
        results.append(result)

    message = "结构变更预览完成" if preview else "结构变更执行完成"
    return _ok(results, message=message)


@require_GET
def schema_sync_page(request):
    """前端页面: 结构同步定时任务管理。"""
    return render(request, "common/schema_sync.html", {})


@require_GET
def schema_sync_task_detail(request):
    task = load_task()
    return _ok({**task, "cron": cron_state(task)})


@csrf_exempt
@require_POST
def schema_sync_task_save(request):
    """保存结构同步定时任务配置, 并同步写入 crontab。"""
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)

    task = load_task()
    for key in ("database", "tables", "doris_database"):
        if payload.get(key) not in (None, ""):
            task[key] = str(payload[key]).strip()
    if "apply" in payload:
        task["apply"] = bool(payload["apply"])
    if "enabled" in payload:
        task["enabled"] = bool(payload["enabled"])
    try:
        minute = int(payload.get("cron_minute", task["cron_minute"]))
        hour = int(payload.get("cron_hour", task["cron_hour"]))
        if not (0 <= minute <= 59 and 0 <= hour <= 23):
            raise ValueError("cron 时间不合法")
        task["cron_minute"] = minute
        task["cron_hour"] = hour
    except (ValueError, TypeError):
        return _fail("cron 时间不合法(minute 0-59, hour 0-23)", code=400, status=400)

    if not task["database"] or not task["tables"]:
        return _fail("database 和 tables 不能为空", code=400, status=400)

    task = save_task(task)
    try:
        if task["enabled"]:
            cron = install_cron(task)
        else:
            cron = remove_cron()
    except RuntimeError as exc:
        return _fail(str(exc), code=500, status=500)
    return _ok(
        {"task": task, "cron": {**cron, **cron_state(task)}},
        message="任务配置已保存",
    )


@csrf_exempt
@require_POST
def schema_sync_run_now(request):
    """按任务配置立即执行/预览结构同步。body: {preview: true/false}"""
    payload, err = _parse_json_body(request)
    preview = bool((payload or {}).get("preview", False))
    task = load_task()
    mysql_config = get_database_config({"db_type": "mysql", "database": task["database"]})
    doris_config = get_doris_config({"doris_database": task["doris_database"]})
    tables = [t.strip() for t in task["tables"].split(",") if t.strip()]

    results = []
    for table in tables:
        try:
            result = sync_table_schema(
                mysql_config,
                doris_config,
                task["database"],
                table,
                doris_database=task["doris_database"],
                preview=preview,
                drop_columns=True,
                auto_create=True,
            )
        except Exception as exc:
            result = {"table": table, "error": str(exc)}
        results.append(result)
    return _ok(results, message="预览完成" if preview else "执行完成")


@require_GET
def schema_sync_log_view(request):
    logs = sorted(
        (PROJECT_ROOT / "logs").glob("schema_sync_*.log"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not logs:
        return _ok({"file": None, "tail": "暂无日志"})
    latest = logs[0]
    tail = "\n".join(
        latest.read_text(encoding="utf-8", errors="replace").splitlines()[-100:]
    )
    return _ok({"file": latest.name, "tail": tail})


@require_GET
def datax_page(request):
    """前端页面: DataX 同步管理。"""
    return render(request, "common/datax.html", {})


@require_GET
def etl_page(request):
    """前端页面: ETL 管理。"""
    return render(request, "common/etl.html", {})


def _read_etl_config() -> dict:
    path = PROJECT_ROOT / "etl" / "config.json"
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return {"kafka_bootstrap_servers": "192.168.3.100:9092",
            "kafka_group_id": "etl_doris_pg_debezium_t1",
            "kafka_topics": ["cdcpg.public.orders"],
            "doris_database": "test_db"}


@require_GET
def etl_config_view(request):
    config = _read_etl_config()
    return _ok(
        {
            "kafka_bootstrap_servers": config.get("kafka_bootstrap_servers", ""),
            "kafka_group_id": config.get("kafka_group_id", ""),
            "doris_database": config.get("doris_database", ""),
            "kafka_topics": config.get("kafka_topics", []),
        }
    )


@csrf_exempt
@require_POST
def etl_config_save(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    path = PROJECT_ROOT / "etl" / "config.json"
    config = _read_etl_config()

    if payload.get("kafka_bootstrap_servers"):
        config["kafka_bootstrap_servers"] = str(payload["kafka_bootstrap_servers"]).strip()
    if payload.get("kafka_group_id"):
        config["kafka_group_id"] = str(payload["kafka_group_id"]).strip()
    if payload.get("doris_database"):
        config["doris_database"] = str(payload["doris_database"]).strip()
    if payload.get("kafka_topics") is not None:
        raw = payload["kafka_topics"]
        if isinstance(raw, str):
            topics = [t.strip() for t in raw.replace("\n", ",").split(",") if t.strip()]
        elif isinstance(raw, list):
            topics = [str(t).strip() for t in raw if str(t).strip()]
        else:
            return _fail("kafka_topics 格式不正确", code=400, status=400)
        if not topics:
            return _fail("kafka_topics 不能为空", code=400, status=400)
        config["kafka_topics"] = topics

    path.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return _ok(
        {
            "kafka_bootstrap_servers": config["kafka_bootstrap_servers"],
            "kafka_group_id": config["kafka_group_id"],
            "doris_database": config["doris_database"],
            "kafka_topics": config["kafka_topics"],
        },
        message="ETL 配置已保存",
    )


@csrf_exempt
@require_POST
def etl_run(request):
    """后台启动 ETL(立即返回), 输出写入 logs/etl_<日期>.log。"""
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    date = str(payload.get("date") or "").strip()
    if not date:
        import datetime as _dt
        date = (_dt.datetime.now() - _dt.timedelta(days=1)).strftime("%F")
    dry_run = bool(payload.get("dry_run", False))

    log_dir = PROJECT_ROOT / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / f"etl_{date}.log"

    command = [
        sys.executable,
        str(PROJECT_ROOT / "etl" / "etl_kafka_doris.py"),
        "--date",
        date,
    ]
    if dry_run:
        command.append("--dry-run")
    with open(log_file, "ab") as fp:
        process = subprocess.Popen(
            command,
            cwd=str(PROJECT_ROOT),
            stdout=fp,
            stderr=subprocess.STDOUT,
        )
    return _ok(
        {"started": True, "pid": process.pid, "date": date, "dry_run": dry_run, "log_file": str(log_file)},
        message=f"ETL 已启动 (pid={process.pid}), 日志: {log_file.name}",
    )


@require_GET
def etl_log_view(request):
    logs = sorted(
        (PROJECT_ROOT / "logs").glob("etl_*.log"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not logs:
        return _ok({"file": None, "tail": "暂无日志"})
    latest = logs[0]
    tail = "\n".join(
        latest.read_text(encoding="utf-8", errors="replace").splitlines()[-100:]
    )
    return _ok({"file": latest.name, "tail": tail})


@require_GET
def flink_sql_page(request):
    """前端页面: Flink SQL 作业管理。"""
    return render(request, "common/flink_sql.html", {})


@require_GET
def flink_sql_files(request):
    files = []
    for path in sorted(FLINK_SQL_DIR.glob("*.sql")):
        stat = path.stat()
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        files.append(
            {
                "name": path.name,
                "size": stat.st_size,
                "mtime": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "preview": "\n".join(lines[:20]),
            }
        )
    return _ok(files)


@require_GET
def flink_sql_file(request):
    name = request.GET.get("name", "")
    if not name or "/" in name or "\\" in name or not name.endswith(".sql"):
        return _fail("文件名不合法", code=400, status=400)
    path = FLINK_SQL_DIR / name
    if not path.exists():
        return _fail("文件不存在", code=404, status=404)
    return _ok(
        {"name": name, "content": path.read_text(encoding="utf-8", errors="replace")}
    )


@require_GET
def flink_sync_jobs(request):
    return _ok(check_all_jobs())


@csrf_exempt
@require_POST
def flink_sync_generate(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    try:
        result = generate_job((payload or {}).get("job"))
    except Exception as exc:
        return _fail(f"生成失败: {exc}", code=500, status=500)
    return _ok(result, message="SQL 已重新生成")


@csrf_exempt
@require_POST
def flink_sync_apply(request):
    """完整流程: Doris 结构同步 -> savepoint 停止 -> 生成 SQL -> 提交。"""
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    job_name = (payload or {}).get("job")
    if not job_name:
        return _fail("缺少 job 参数", code=400, status=400)
    try:
        result = apply_job(job_name, doris_sync=bool((payload or {}).get("doris_sync", True)))
    except Exception as exc:
        return _fail(f"执行失败: {exc}", code=500, status=500)
    return _ok(result, message="变更流程执行完成")


# ---------------------------------------------------------------- 元数据源配置

JDBC_TEMPLATES = {
    "mysql": "jdbc:mysql://{host}:{port}/{database}?useUnicode=true&characterEncoding=utf8",
    "postgresql": "jdbc:postgresql://{host}:{port}/{database}",
    "oracle": "jdbc:oracle:thin:@{host}:{port}:{database}",
    "hive": "jdbc:hive2://{host}:{port}/{database}",
    "doris": "jdbc:mysql://{host}:{port}/{database}",
    "sqlserver": "jdbc:sqlserver://{host}:{port};databaseName={database}",
    "kafka": "kafka://{host}:{port}",
    "other": "",
}

DEFAULT_PORTS = {
    "mysql": 3306,
    "postgresql": 5432,
    "oracle": 1521,
    "hive": 10000,
    "doris": 9030,
    "sqlserver": 1433,
    "kafka": 9092,
}


def _build_jdbc_url(db_type: str, host: str, port, database: str) -> str:
    template = JDBC_TEMPLATES.get(db_type, "")
    if not template:
        return ""
    return template.format(
        host=host or "",
        port=port or DEFAULT_PORTS.get(db_type, ""),
        database=database or "",
    )


def _fill_source_from_payload(source, payload: dict, creating: bool = False) -> None:
    if payload.get("name") not in (None, ""):
        source.name = str(payload["name"]).strip()
    if payload.get("db_type") in SourceType.values:
        source.db_type = payload["db_type"]
    if payload.get("host") is not None:
        source.host = str(payload["host"]).strip()
    if payload.get("port") not in (None, ""):
        try:
            source.port = int(payload["port"])
        except (TypeError, ValueError):
            raise ValueError("port 必须是数字")
    if payload.get("database_name") is not None:
        source.database_name = str(payload["database_name"]).strip()
    if payload.get("schema_name") is not None:
        source.schema_name = str(payload["schema_name"]).strip()
    if payload.get("username") is not None:
        source.username = str(payload["username"]).strip()
    if payload.get("password") not in (None, ""):
        source.password = str(payload["password"])
    if payload.get("remark") is not None:
        source.remark = str(payload["remark"]).strip()
    if "enabled" in payload:
        source.enabled = bool(payload["enabled"])
    raw_jdbc = payload.get("jdbc_url")
    if raw_jdbc not in (None, ""):
        source.jdbc_url = str(raw_jdbc).strip()
    elif creating or not source.jdbc_url:
        source.jdbc_url = _build_jdbc_url(
            source.db_type, source.host, source.port, source.database_name
        )
    if source.port in (None, 0, ""):
        source.port = DEFAULT_PORTS.get(source.db_type)


@require_GET
def sources_page(request):
    """前端页面: 元数据源配置。"""
    return render(
        request,
        "common/sources.html",
        {
            "source_types": SourceType.choices,
            "jdbc_templates": JDBC_TEMPLATES,
            "default_ports": DEFAULT_PORTS,
        },
    )


@require_GET
def source_list(request):
    sources = MetadataSourceConfig.objects.all()
    return _ok([source_config_to_dict(s) for s in sources])


@csrf_exempt
@require_POST
def source_create(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    name = str(payload.get("name") or "").strip()
    if not name:
        return _fail("name 不能为空", code=400, status=400)
    if MetadataSourceConfig.objects.filter(name=name).exists():
        return _fail(f"名称已存在: {name}", code=400, status=400)
    source = MetadataSourceConfig(name=name)
    try:
        _fill_source_from_payload(source, payload, creating=True)
    except ValueError as exc:
        return _fail(str(exc), code=400, status=400)
    source.save()
    return _ok(source_config_to_dict(source), message="配置已创建")


@csrf_exempt
@require_POST
def source_update(request, pk):
    source = get_object_or_404(MetadataSourceConfig, pk=pk)
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    if payload.get("name") and MetadataSourceConfig.objects.filter(
        name=str(payload["name"]).strip()
    ).exclude(pk=source.pk).exists():
        return _fail(f"名称已存在: {payload['name']}", code=400, status=400)
    try:
        _fill_source_from_payload(source, payload)
    except ValueError as exc:
        return _fail(str(exc), code=400, status=400)
    source.save()
    return _ok(source_config_to_dict(source), message="配置已更新")


@csrf_exempt
@require_POST
def source_delete(request, pk):
    source = get_object_or_404(MetadataSourceConfig, pk=pk)
    source.delete()
    return _ok({"id": pk}, message="配置已删除")


@csrf_exempt
@require_POST
def source_test(request, pk):
    """测试连接: mysql/postgres/doris 真连, 其余做 TCP 探测。"""
    source = get_object_or_404(MetadataSourceConfig, pk=pk)
    host = source.host or ""
    port = source.port or DEFAULT_PORTS.get(source.db_type)
    if not host:
        return _fail("未配置 host, 无法测试", code=400, status=400)
    try:
        if source.db_type in ("mysql", "doris"):
            import pymysql
            conn = pymysql.connect(
                host=host,
                port=port or 3306,
                user=source.username or "",
                password=source.password,
                database=source.database_name or None,
                connect_timeout=5,
            )
            conn.close()
            return _ok({"ok": True}, message="连接成功")
        if source.db_type == "postgresql":
            import psycopg2
            conn = psycopg2.connect(
                host=host,
                port=port or 5432,
                dbname=source.database_name or "postgres",
                user=source.username or "",
                password=source.password,
                connect_timeout=5,
            )
            conn.close()
            return _ok({"ok": True}, message="连接成功")
        # 其余类型: TCP 连通性 + 提示
        import socket
        with socket.create_connection((host, port or 0), timeout=5):
            return _ok(
                {"ok": True, "tcp_only": True},
                message=f"TCP 可达; {source.get_db_type_display()} 的 JDBC 驱动测试请在本机安装对应驱动",
            )
    except Exception as exc:
        return _fail(f"连接失败: {exc}", code=500, status=500)


@csrf_exempt
@require_POST
def source_sync_metadata(request, pk):
    """用配置的连接信息同步元数据(支持 mysql / postgresql)。"""
    source = get_object_or_404(MetadataSourceConfig, pk=pk)
    if source.db_type not in ("mysql", "postgresql"):
        return _fail(f"{source.db_type} 暂不支持元数据自动同步", code=400, status=400)
    config = {
        "db_type": source.db_type,
        "host": source.host,
        "port": source.port or DEFAULT_PORTS.get(source.db_type),
        "user": source.username,
        "password": source.password,
        "database": source.database_name,
        "schema": source.schema_name or None,
        "name": source.name,
    }
    try:
        database, stats = sync_metadata(config)
    except Exception as exc:
        return _fail(f"同步失败: {exc}", code=500, status=500)
    return _ok(
        {"database": database_to_dict(database), "stats": stats},
        message=f"同步完成: 表 {stats['tables']}, 字段 {stats['columns']}",
    )


# ---------------------------------------------------------------- SQL 助手

@require_GET
def sql_helper_page(request):
    """前端页面: SQL 助手。"""
    return render(request, "common/sql_helper.html", {})


@require_GET
def sql_helper_options(request):
    databases = MetadataDatabase.objects.all()
    return _ok(
        [
            {
                "id": db.id,
                "name": db.name or f"{db.db_type}://{db.host}",
                "db_type": db.db_type,
                "host": db.host,
                "database_name": db.database_name,
            }
            for db in databases
        ]
    )


@require_GET
def sql_helper_tables(request):
    db_id = request.GET.get("db_id")
    tables = MetadataTable.objects.none()
    if db_id:
        tables = (
            MetadataTable.objects.filter(database_id=db_id)
            .order_by("schema_name", "name")
        )
    return _ok(
        [
            {"id": t.id, "schema_name": t.schema_name, "name": t.name, "comment": t.comment}
            for t in tables
        ]
    )


@require_GET
def sql_helper_table(request, pk):
    table = get_object_or_404(
        MetadataTable.objects.select_related("database").prefetch_related("columns", "indexes"),
        pk=pk,
    )
    columns = list(table.columns.all())
    snippets = build_snippets(table, columns, table.database)
    return _ok(
        {
            "table": {
                "id": table.id,
                "schema_name": table.schema_name,
                "name": table.name,
                "comment": table.comment,
            },
            "database": {
                "id": table.database_id,
                "db_type": table.database.db_type,
                "database_name": table.database.database_name,
            },
            "columns": [
                {
                    "name": c.name,
                    "data_type": c.data_type,
                    "column_type": c.column_type,
                    "is_nullable": c.is_nullable,
                    "max_length": c.max_length,
                    "comment": c.comment,
                }
                for c in columns
            ],
            "snippets": snippets,
        }
    )


# ---------------------------------------------------------------- 对账中心

@require_GET
def reconcile_page(request):
    return render(
        request,
        "common/reconcile.html",
        {"task_types": ReconcileTask._meta.get_field("task_type").choices},
    )


def _reconcile_task_to_dict(task: ReconcileTask) -> dict:
    return {
        "id": task.id,
        "name": task.name,
        "task_type": task.task_type,
        "task_type_label": task.get_task_type_display(),
        "source_config_id": task.source_config_id,
        "source_config_name": task.source_config.name if task.source_config else "",
        "source_db_name": task.source_db_name,
        "source_schema": task.source_schema,
        "target_db_name": task.target_db_name,
        "tables": task.tables,
        "columns": task.columns,
        "pk_columns": task.pk_columns,
        "metric_sql": task.metric_sql,
        "enabled": task.enabled,
        "remark": task.remark,
        "updated_at": task.updated_at.isoformat(),
        "run_count": task.runs.count(),
        "last_run": (
            {
                "id": task.runs.first().id,
                "status": task.runs.first().status,
                "summary": task.runs.first().summary,
                "ran_at": task.runs.first().ran_at.isoformat(),
            }
            if task.runs.exists()
            else None
        ),
    }


@require_GET
def reconcile_task_list(request):
    tasks = ReconcileTask.objects.prefetch_related("runs").order_by("-updated_at")
    return _ok([_reconcile_task_to_dict(t) for t in tasks])


@csrf_exempt
@require_POST
def reconcile_task_create(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    name = str(payload.get("name") or "").strip()
    if not name:
        return _fail("name 不能为空", code=400, status=400)
    task = ReconcileTask(name=name)
    message = _apply_reconcile_payload(task, payload)
    if message:
        return _fail(message, code=400, status=400)
    task.save()
    return _ok(_reconcile_task_to_dict(task), message="任务已创建")


@csrf_exempt
@require_POST
def reconcile_task_update(request, pk):
    task = get_object_or_404(ReconcileTask, pk=pk)
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    if payload.get("name") not in (None, ""):
        task.name = str(payload["name"]).strip()
    message = _apply_reconcile_payload(task, payload)
    if message:
        return _fail(message, code=400, status=400)
    task.save()
    return _ok(_reconcile_task_to_dict(task), message="任务已更新")


def _apply_reconcile_payload(task: ReconcileTask, payload: dict) -> str | None:
    task_type = payload.get("task_type")
    if task_type not in dict(ReconcileTask._meta.get_field("task_type").choices):
        return "task_type 不合法"
    task.task_type = task_type
    if "source_config_id" in payload:
        task.source_config_id = payload.get("source_config_id") or None
    for key in ("source_db_name", "source_schema", "target_db_name", "metric_sql", "remark"):
        if payload.get(key) is not None:
            setattr(task, key, str(payload[key]).strip())
    for key in ("tables", "columns", "pk_columns"):
        if payload.get(key) is not None:
            setattr(
                task,
                key,
                [str(item).strip() for item in payload[key] if str(item).strip()],
            )
    if "enabled" in payload:
        task.enabled = bool(payload["enabled"])
    if not task.tables:
        return "tables 不能为空"
    if task.task_type == "pk_snapshot" and not task.pk_columns:
        return "主键快照对账需要配置 pk_columns"
    if task.task_type == "metric" and not task.metric_sql:
        return "业务指标对账需要配置 metric_sql"
    return None


@require_GET
def reconcile_task_detail(request, pk):
    task = get_object_or_404(
        ReconcileTask.objects.prefetch_related("runs"), pk=pk
    )
    runs = [
        {
            "id": r.id,
            "status": r.status,
            "summary": r.summary,
            "error": r.error,
            "duration_ms": r.duration_ms,
            "ran_at": r.ran_at.isoformat(),
            "details": r.details,
        }
        for r in task.runs.all()[:10]
    ]
    return _ok({"task": _reconcile_task_to_dict(task), "runs": runs})


@csrf_exempt
@require_POST
def reconcile_task_delete(request, pk):
    task = get_object_or_404(ReconcileTask, pk=pk)
    task.delete()
    return _ok({"id": pk}, message="任务已删除")


@csrf_exempt
@require_POST
def reconcile_task_run(request, pk):
    task = get_object_or_404(
        ReconcileTask.objects.select_related("source_config"), pk=pk
    )
    if not task.source_config:
        return _fail("任务未关联源配置", code=400, status=400)
    run = run_reconcile_task(task)
    return _ok(
        {
            "run_id": run.id,
            "status": run.status,
            "summary": run.summary,
            "error": run.error,
            "duration_ms": run.duration_ms,
            "details": run.details,
        },
        message="对账完成" if run.status == "success" else "对账失败",
    )


# ---------------------------------------------------------------- 文档 Web

@require_GET
def docs_page(request):
    return render(request, "common/docs.html", {})


@require_GET
def docs_list(request):
    return _ok(docs_service.list_docs())


@require_GET
def docs_detail(request):
    name = request.GET.get("name", "")
    try:
        return _ok(docs_service.get_doc(name))
    except (ValueError, FileNotFoundError) as exc:
        return _fail(str(exc), code=404, status=404)


# ---------------------------------------------------------------- SQL 文件库

@require_GET
def sql_files_page(request):
    return render(request, "common/sql_files.html", {})


@require_GET
def sql_files_list(request):
    try:
        entries = sql_files_service.backend().list_files(request.GET.get("path", ""))
    except Exception as exc:
        return _fail(f"读取失败: {exc}", code=500, status=500)
    return _ok(
        {
            "backend": sql_files_service.backend().label,
            "base_dir": str(sql_files_service.base_dir()),
            "entries": entries,
        }
    )


@require_GET
def sql_files_read(request):
    path = request.GET.get("path", "")
    try:
        content = sql_files_service.backend().read_file(path)
    except Exception as exc:
        return _fail(f"读取失败: {exc}", code=500, status=500)
    return _ok({"path": path, "content": content})


# ---------------------------------------------------------------- 血缘

@require_GET
def lineage_page(request):
    return render(request, "common/lineage.html", {})


@require_GET
def lineage_graph(request):
    return _ok(lineage_service.graph())


@csrf_exempt
@require_POST
def lineage_parse(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    sql_text = payload.get("sql") or ""
    sql_file = str(payload.get("sql_file") or "").strip()
    parsed = lineage_service.parse_sql(sql_text)
    saved = []
    if payload.get("save") and parsed:
        saved = lineage_service.save_lineage(sql_file, sql_text)
    return _ok({"parsed": parsed, "saved": saved}, message="解析完成")


@csrf_exempt
@require_POST
def lineage_clear(request):
    deleted, _ = LineageEdge.objects.all().delete()
    return _ok({"deleted": deleted}, message="血缘已清空")


# ---------------------------------------------------------------- LLM 分析

@csrf_exempt
@require_POST
def llm_analyze(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    kind = payload.get("kind", "sql")
    metadata = ""
    if payload.get("table_id"):
        table = MetadataTable.objects.filter(pk=payload["table_id"]).select_related("database").first()
        if table:
            columns = [
                f"{c.name} {c.column_type or c.data_type}"
                + (f" 注释:{c.comment}" if c.comment else "")
                for c in table.columns.all()
            ]
            metadata = (
                f"表 {table.database.database_name}.{table.name} 注释:{table.comment}\n"
                + "\n".join(columns)
            )
    elif payload.get("metadata"):
        metadata = str(payload["metadata"])
    try:
        if kind == "sql":
            text = llm_service.analyze_sql(str(payload.get("sql") or ""), metadata)
        else:
            text = llm_service.analyze_metadata(metadata or str(payload.get("sql") or ""))
    except llm_service.LLMNotConfigured as exc:
        return _fail(str(exc), code=400, status=400)
    except Exception as exc:
        return _fail(f"分析失败: {exc}", code=500, status=500)
    return _ok({"analysis": text}, message="分析完成")


@require_GET
def ai_sql_page(request):
    """前端页面: AI 辅助写 SQL。"""
    return render(request, "common/ai_sql.html", {})


@csrf_exempt
@require_POST
def llm_sql_assist(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    request_text = str(payload.get("request") or "").strip()
    mode = str(payload.get("mode") or "generate").strip()
    if not request_text:
        return _fail("request 不能为空", code=400, status=400)
    if mode not in ("generate", "optimize", "explain"):
        return _fail("mode 只能是 generate/optimize/explain", code=400, status=400)

    metadata = ""
    if payload.get("table_id"):
        table = (
            MetadataTable.objects.filter(pk=payload["table_id"])
            .select_related("database")
            .first()
        )
        if table:
            columns = [
                f"{c.name} {c.column_type or c.data_type}"
                + (f" 注释:{c.comment}" if c.comment else "")
                for c in table.columns.all()
            ]
            metadata = (
                f"表 {table.database.database_name}.{table.name} 注释:{table.comment}\n"
                + "\n".join(columns)
            )
    elif payload.get("metadata"):
        metadata = str(payload["metadata"])

    try:
        text = llm_service.sql_assist(request_text, mode, metadata)
    except llm_service.LLMNotConfigured as exc:
        return _fail(str(exc), code=400, status=400)
    except Exception as exc:
        return _fail(f"AI 调用失败: {exc}", code=500, status=500)
    return _ok({"mode": mode, "result": text}, message="AI 完成")


# ---------------------------------------------------------------- 运营看板

@require_GET
def ops_page(request):
    """前端页面: 运营看板。"""
    return render(request, "common/ops.html", {})


@require_GET
def ops_summary(request):
    try:
        days = max(1, min(int(request.GET.get("days", 7)), 90))
    except (TypeError, ValueError):
        days = 7
    from datetime import timedelta

    since = timezone.now() - timedelta(days=days)
    base = AnalyticsEvent.objects.filter(created_at__gte=since)
    total = base.count()
    errors = base.filter(status_code__gte=400).count()
    by_day = list(
        base.annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(count=Count("id"))
        .order_by("day")
    )
    top_paths = list(
        base.values("method", "path")
        .annotate(count=Count("id"))
        .order_by("-count")[:12]
    )
    status_dist = list(base.values("status_code").annotate(count=Count("id")).order_by("status_code"))
    recent = list(
        base.order_by("-created_at")[:30].values(
            "created_at", "method", "path", "status_code", "duration_ms", "username"
        )
    )
    return _ok(
        {
            "days": days,
            "total": total,
            "errors": errors,
            "error_rate": round(errors / total * 100, 2) if total else 0,
            "success_rate": round((total - errors) / total * 100, 2) if total else 0,
            "by_day": [
                {"day": item["day"].isoformat(), "count": item["count"]}
                for item in by_day
            ],
            "top_paths": top_paths,
            "status_dist": status_dist,
            "recent": recent,
        }
    )


# ---------------------------------------------------------------- 脚本管理

@require_GET
def scripts_page(request):
    return render(request, "common/scripts.html", {})


@require_GET
def scripts_list(request):
    try:
        entries = scripts_service.list_scripts()
    except Exception as exc:
        return _fail(f"读取失败: {exc}", code=500, status=500)
    return _ok(
        {
            "dirs": scripts_service.SCRIPT_DIRS,
            "scripts": entries,
        }
    )


@require_GET
def scripts_read(request):
    path = request.GET.get("path", "")
    try:
        content = scripts_service.read_script(path)
    except Exception as exc:
        return _fail(f"读取失败: {exc}", code=500, status=500)
    return _ok({"path": path, "content": content})


@csrf_exempt
@require_POST
def scripts_save(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    path = str(payload.get("path") or "").strip()
    content = str(payload.get("content") or "")
    try:
        result = scripts_service.save_script(path, content)
    except Exception as exc:
        return _fail(f"保存失败: {exc}", code=500, status=500)
    return _ok(result, message="脚本已保存")


@csrf_exempt
@require_POST
def scripts_create(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    try:
        result = scripts_service.create_script(
            str(payload.get("directory") or "tools").strip(),
            str(payload.get("name") or "").strip(),
            str(payload.get("content") or ""),
        )
    except Exception as exc:
        return _fail(f"创建失败: {exc}", code=500, status=500)
    return _ok(result, message="脚本已创建")


@csrf_exempt
@require_POST
def scripts_delete(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    try:
        result = scripts_service.delete_script(str(payload.get("path") or "").strip())
    except Exception as exc:
        return _fail(f"删除失败: {exc}", code=500, status=500)
    return _ok(result, message="脚本已删除")


@csrf_exempt
@require_POST
def scripts_run(request):
    payload, err = _parse_json_body(request)
    if err:
        return _fail(err, code=400, status=400)
    try:
        result = scripts_service.run_script(
            str(payload.get("path") or "").strip(),
            args=[str(a) for a in (payload.get("args") or [])],
            timeout=int(payload.get("timeout", 600)),
        )
    except Exception as exc:
        return _fail(f"运行失败: {exc}", code=500, status=500)
    message = {
        "success": "执行完成",
        "failed": "执行失败",
        "timeout": "执行超时",
    }.get(result["status"], "执行完成")
    return _ok(result, message=message)


@require_GET
def scripts_runs(request):
    from .models import ScriptRun

    runs = list(
        ScriptRun.objects.order_by("-started_at")[:30].values(
            "id", "script_path", "args", "status", "exit_code", "duration_ms", "started_at"
        )
    )
    return _ok(runs)
